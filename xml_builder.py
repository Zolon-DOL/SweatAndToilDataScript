# coding: utf-8

from openpyxl import load_workbook
import xml.etree.ElementTree as elementTree
import re
import textwrap

COUNTRIES_OUTPUT_FILE = 'countries_output.xml'
GOODS_OUTPUT_FILE = 'goods_output.xml'
#wb = load_workbook('master_data-1.xlsx')
#wb = load_workbook('master_data_updated_8.16.19.xlsx')
wb = load_workbook('master_data_updated_8.20.19.xlsx')

countries = elementTree.Element('Countries')
countries_tree = elementTree.ElementTree(countries)

goods = elementTree.Element('Goods')
goods_tree = elementTree.ElementTree(goods)

year = '2018'

# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
country_display_names = {
        'RS': 'Republika Srpska',
        'FBiH': 'Federation of Bosnia and Herzegovina',
        'BD': 'Br\u+010Dko District',  # 'Brčko District'  'Br\u+010Dko District'
        'BiH': 'Bosnia and Herzegovina'
    }

# 'Br\xc4\x8dko District'
multi_territory_countries = ['Tanzania', 'Pakistan', 'Somalia', 'Iraq', 'Bosnia and Herzegovina']

# Matching cases such as:
# hello(123)world, hello( 123)world, hello (1,23 )world, hello( 12 , 3 )world
# hello*(123)world, hello*( 123)world, hello *(1,23 )world, hello*( 12 , 3 )world
parentheses = re.compile(r"\s*\*?\(\s*?\d+\s*?,*\s*?\d*\s*?\)")


def format_text(phrase=''):
        return textwrap.fill(text=phrase)


def transform_data_to_list(row, apply_parentheses_regex=0):
    row_to_list = []
    for column_index, value in enumerate(row):
        if isinstance(value, long):
            value_as_string = str(it_row[column_index].real)
            row_to_list.append(value_as_string)
        else:
            if apply_parentheses_regex and column_index == 5:
                if value is not None and re.findall(parentheses, value):
                    value = value.rstrip()
                    matches = re.finditer(parentheses, value)
                    for indices in matches:
                        if indices.regs[0][1] >= len(value) - 1:
                            value = re.sub(parentheses, '', value).rstrip()

            row_to_list.append(value)

    return row_to_list


def country_exists(country_name):
    for country in countries.findall('Country'):
        name = country.find('Name').text
        if name == country_name:
            return country
    return None


def check_multiple_territories(country):
    territories = country.find('Multiple_Territories')
    if territories is None:
        territories = elementTree.SubElement(country, 'Multiple_Territories')

    name = country.find('Name').text

    if name in multi_territory_countries:
        territories.text = 'Yes'
    else:
        territories.text = 'No'


def country_profiles(country, row):
    data = transform_data_to_list(row)

    region = data[2]
    advancement_lvl = data[3]
    description = data[4].replace('\n', ' ')

    elementTree.SubElement(country, 'Region').text = region
    elementTree.SubElement(country, 'Multiple_Territories')
    elementTree.SubElement(country, 'Advancement_Level').text = advancement_lvl
    elementTree.SubElement(country, 'Description').text = description

    # create these tags now to be used in later sheets
    elementTree.SubElement(country, 'Goods')


def goods_list(country, row):
    data = transform_data_to_list(row)
    good = data[2]

    child_labor, forced_labor, forced_child_labor = 'No', 'No', 'No'
    labor_type = {0: child_labor, 1: forced_labor, 2: forced_child_labor}
    for labor_index, value in enumerate(data[3:6]):
        if value == 1:
            labor_type[labor_index] = 'Yes'
    child_labor = 'Yes' if data[3] == 1 else 'No'
    forced_labor = 'Yes' if data[4] == 1 else 'No'
    forced_child_labor = 'Yes' if data[5] == 1 else 'No'

    sectors = {'manu': 'Manufacturing',
               'mine': 'Mining',
               'agri': 'Agriculture',
               'other': 'Other'}
    sector = sectors[data[6]] if data[6] in sectors else ''
    if sector:
        # countries.xml
        goods_tag = country.find('Goods')
        if goods_tag is None:
            goods_tag = elementTree.SubElement(country, 'Goods')

        good_tag = elementTree.SubElement(goods_tag, 'Good')
        elementTree.SubElement(good_tag, 'Good_Name').text = good
        elementTree.SubElement(good_tag, 'Child_Labor').text = child_labor
        elementTree.SubElement(good_tag, 'Forced_Labor').text = forced_labor
        elementTree.SubElement(good_tag, 'Forced_Child_Labor').text = forced_child_labor

        # goods.xml
        good_tag = None
        for val in goods.findall('Good'):
            name = val.find('Good_Name')
            if name.text == good:
                good_tag = val
                break

        if good_tag is None:
            good_tag = elementTree.SubElement(goods, 'Good')
            elementTree.SubElement(good_tag, 'Good_Name').text = good
            elementTree.SubElement(good_tag, 'Good_Sector').text = sector

        countries_tag = good_tag.find('Countries')
        if countries_tag is None:
            countries_tag = elementTree.SubElement(good_tag, 'Countries')
        country_tag = elementTree.SubElement(countries_tag, 'Country')
        country_name = country.find('Name')
        country_region = country.find('Region')

        # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
        # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
        # ET.SubElement(countryTag, "Country_Name").text = countryName.text if not countryName == None else ""
        # ET.SubElement(country_tag, 'Country_Name').text = country_name.text if country_name is None else ''
        if country.find('Name') is not None and country.find('Name').text is not None:
            elementTree.SubElement(country_tag, 'Country_Name').text = country_name.text
        else:
            elementTree.SubElement(country_tag, 'Country_Name').text = ''

        # ET.SubElement(countryTag, "Country_Region").text = countryRegion.text if not countryRegion == None else ""
        # ET.SubElement(country_tag, 'Country_Region').text = country_region.text if country_region is None else ''
        if country.find('Region') is not None and country.find('Region').text is not None:
            elementTree.SubElement(country_tag, 'Country_Region').text = country_region.text
        else:
            elementTree.SubElement(country_tag, 'Country_Region').text = ''

        elementTree.SubElement(country_tag, 'Child_Labor').text = child_labor
        elementTree.SubElement(country_tag, 'Forced_Labor').text = forced_labor
        elementTree.SubElement(
            country_tag, 'Forced_Child_Labor').text = forced_child_labor


def statistics_on_children(country, row):
    stats = country.find('Country_Statistics')
    if stats is None:
        stats = elementTree.SubElement(country, 'Country_Statistics')

    regex = r"(^\d+(,\d+)*(\.\d+(e\d+)?)?)(\s\((\d+(,\d+)*(\.\d+(e\d+)?)?)\))?$"

    data = transform_data_to_list(row)

    stat_type = data[3]
    sector = data[4]
    age = data[5]
    percent = data[6]
    match = re.match(regex, str(percent))

    age_range = age.replace('to', '-').replace(' ', '') if age else ''
    group = match.group(1) if match else ''
    percentage = str(round(float(group) / 100, 3)) if is_number(group) else 'Unavailable'

    check_multiple_territories(country)

    if stat_type == 'Working (% and population)' or stat_type == 'Working children by sector':
        child_work = stats.find('Children_Work_Statistics')
        if child_work is None:
            child_work = elementTree.SubElement(stats, 'Children_Work_Statistics')

        if stat_type == 'Working (% and population)':
            total_work_pop = match.group(6) if match else ''
            if total_work_pop:
                total_work_pop = total_work_pop.replace(',', '')

            elementTree.SubElement(child_work, 'Age_Range').text = age_range
            elementTree.SubElement(
                child_work, 'Total_Percentage_of_Working_Children').text = percentage
            elementTree.SubElement(
                child_work, 'Total_Working_Population').text = total_work_pop
        elif stat_type == 'Working children by sector' and sector:
            elementTree.SubElement(child_work, sector).text = percentage
    elif stat_type == 'Attending School (%)':
        education = stats.find('Education_Statistics_Attendance_Statistics')
        if education is None:
            education = elementTree.SubElement(
                stats, 'Education_Statistics_Attendance_Statistics')

        elementTree.SubElement(education, 'Age_Range').text = age_range
        elementTree.SubElement(
            education, 'Percentage').text = percentage
    elif stat_type == 'Combining Work and School (%)':
        work_and_school = stats.find(
            'Children_Working_and_Studying_7-14_yrs_old')
        if work_and_school is None:
            work_and_school = elementTree.SubElement(
                stats, 'Children_Working_and_Studying_7-14_yrs_old')

        elementTree.SubElement(work_and_school, 'Age_Range').text = age_range
        elementTree.SubElement(
            work_and_school, 'Total').text = percentage
    elif stat_type == 'Primary Completion Rate (%)':
        completion_rate = stats.find(
            'UNESCO_Primary_Completion_Rate')
        if completion_rate is None:
            completion_rate = elementTree.SubElement(
                stats, 'UNESCO_Primary_Completion_Rate')
            elementTree.SubElement(
                completion_rate, 'Rate').text = percentage


def ratification_of_international(country, row):
    conventions = country.find('Conventions')
    if conventions is None:
        conventions = elementTree.SubElement(country, 'Conventions')

    data = transform_data_to_list(row)

    convention = data[3]
    ratification = data[4]

    tags = {
        'ILO C. 138, Minimum Age': 'C_138_Ratified',
        'UN CRC': 'Convention_on_the_Rights_of_the_Child_Ratified',
        'ILO C. 182, Worst Forms of Child Labor': 'C_182_Ratified',
        'UN CRC Optional Protocol on the Sale of Children, Child Prostitution and Child Pornography':
            'CRC_Commercial_Sexual_Exploitation_of_Children_Ratified',
        'UN CRC Optional Protocol on Armed Conflict': 'CRC_Armed_Conflict_Ratified',
        'Palermo Protocol on Trafficking in Persons': 'Palermo_Ratified'
    }

    if ratification == '1':
        ratification = 'Yes'
    elif ratification == '0':
        ratification = 'No'
    if convention:
        elementTree.SubElement(conventions, tags[convention]).text = ratification


def laws_and_regulations(country, row):
    legal = country.find('Legal_Standards')
    if legal is None:
        legal = elementTree.SubElement(country, 'Legal_Standards')

    data = transform_data_to_list(row)

    related_entity = data[2]
    standard = data[3]
    meets_intl_stds = data[5]
    age = data[7]
    calced_age = 'Yes' if data[8] == 'TRUE' else 'No'

    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
    # multiple_territories = True if country.find('Multiple_Territories').text == 'Yes' else False
    multiple_territories = False
    if country.find('Multiple_Territories') is not None and country.find('Multiple_Territories').text == 'Yes':
        multiple_territories = True

    tags = {
        'Compulsory Education Age': 'Compulsory_Education',
        'Free Public Education': 'Free_Public_Education',
        'Identification of Hazardous Occupations or Activities Prohibited for Children': 'Types_Hazardous_Work',
        'Minimum Age for Hazardous Work': 'Minimum_Hazardous_Work',
        'Minimum Age for Voluntary State Military Recruitment': 'Minumum_Voluntary_Military',
        'Minimum Age for Work': 'Minimum_Work',
        'Prohibition of Child Trafficking': 'Prohibition_Child_Trafficking',
        'Prohibition of Commercial Sexual Exploitation of Children': 'Prohibition_CSEC',
        'Prohibition of Compulsory Recruitment of Children by (State) Military': 'Minimum_Compulsory_Military',
        'Prohibition of Forced Labor': 'Prohibition_Forced_Labor',
        'Prohibition of Military Recruitment by Non-state Armed Groups': 'Minumum_Non_State_Military',
        'Prohibition of Using Children in Illicit Activities': 'Prohibition_Illicit_Activities'
    }

    if standard and standard in tags:
        tag = legal.find(tags[standard])
        if tag is None:
            tag = elementTree.SubElement(legal, tags[standard])
        if multiple_territories:
            tag = elementTree.SubElement(tag, 'Territory')
            display_name = 'All Territories'
            if related_entity:
                display_name = country_display_names[related_entity] \
                    if related_entity in country_display_names else related_entity
            elementTree.SubElement(tag, 'Territory_Name').text = display_name
            elementTree.SubElement(tag, 'Territory_Display_Name').text = display_name
        elementTree.SubElement(tag, 'Standard').text = meets_intl_stds
        elementTree.SubElement(tag, 'Age').text = age
        elementTree.SubElement(tag, 'Calculated_Age').text = calced_age
        elementTree.SubElement(tag, 'Conforms_To_Intl_Standard').text = meets_intl_stds


def labor_law_enforcement(country, row):
    enforcements = country.find('Enforcements')
    if enforcements is None:
        enforcements = elementTree.SubElement(country, 'Enforcements')

    data = transform_data_to_list(row, 1)

    related_entity = data[2]
    overview = data[3]
    sub_overview = data[4]
    current_year_data = data[5]

    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
    # multiple_territories = True if country.find('Multiple_Territories').text == 'Yes' else False
    multiple_territories = False
    if country.find('Multiple_Territories') is not None and country.find('Multiple_Territories').text == 'Yes':
        multiple_territories = True

    tags = {
        'Labor Inspectorate Funding': 'Labor_Funding',
        'Number of Labor Inspectors': 'Labor_Inspectors',
        'Inspectorate Authorized to Assess Penalties': 'Authorized_Access_Penalties',
        'Initial Training for New Labor Inspectors':
            {
                'NA': 'Labor_New_Employee_Training',
                'Training on New Laws Related to Child Labor': 'Labor_New_Law_Training',
                'Refresher Courses Provided': 'Labor_Refresher_Courses'
            },
        'Number of Labor Inspections Conducted':
            {
                'NA': 'Labor_Inspections',
                'Number Conducted at Worksite': 'Labor_Worksite_Inspections'
            },
        'Number of Child Labor Violations Found':
            {
                'NA': 'Labor_Violations',
                'Number of Child Labor Violations for Which Penalties Were Imposed': 'Labor_Penalties_Imposed',
                'Number of Child Labor Penalties Imposed that Were Collected': 'Labor_Penalties_Collected'
            },
        'Routine Inspections Conducted':
            {
                'NA': 'Labor_Routine_Inspections_Conducted',
                'Routine Inspections Targeted': 'Labor_Routine_Inspections_Targeted'
            },
        'Unannounced Inspections Permitted':
            {
                'NA': 'Labor_Unannounced_Inspections_Premitted',
                'Unannounced Inspections Conducted': 'Labor_Unannounced_Inspections_Conducted'
            },
        'Complaint Mechanism Exists': 'Labor_Complaint_Mechanism',
        'Reciprocal Referral Mechanism Exists Between Labor Authorities and Social Services':
            'Labor_Referral_Mechanism',
        'Number of Labor Inspectors meets ILO Rec': 'Labor_Inspectors_Intl_Standards'
    }

    if overview and (overview in tags and tags[overview]):
        if isinstance(tags[overview], dict):
            if sub_overview is None:
                sub_overview = 'NA'
            tag = enforcements.find(tags[overview][sub_overview])
            if tag is None:
                tag = elementTree.SubElement(
                    enforcements, tags[overview][sub_overview])
        else:
            tag = enforcements.find(tags[overview])
            if tag is None:
                tag = elementTree.SubElement(
                    enforcements, tags[overview])
        if multiple_territories:
            tag = elementTree.SubElement(tag, 'Territory')
            display_name = 'All Territories'
            if related_entity:
                display_name = country_display_names[related_entity] \
                    if related_entity in country_display_names else related_entity
            elementTree.SubElement(tag, 'Territory_Name').text = display_name
            elementTree.SubElement(tag, 'Territory_Display_Name').text = display_name
            elementTree.SubElement(tag, 'Enforcement').text = current_year_data
        else:
            tag.text = current_year_data


def criminal_law_enforcement(country, row):
    enforcements = country.find('Enforcements')
    if enforcements is None:
        enforcements = elementTree.SubElement(country, 'Enforcements')

    data = transform_data_to_list(row, 1)

    related_entity = data[2]
    overview = data[3]
    sub_overview = data[4]
    current_year_data = data[5]

    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
    # multiple_territories = True if country.find('Multiple_Territories').text == 'Yes' else False
    multiple_territories = False
    if country.find('Multiple_Territories') is not None and country.find('Multiple_Territories').text == 'Yes':
        multiple_territories = True

    tags = {
        'Number of Investigations': 'Criminal_Investigations',
        'Number of Violations Found': 'Criminal_Violations',
        'Number of Prosecutions Initiated': 'Criminal_Prosecutions',
        'Number of Convictions': 'Criminal_Convictions',
        'Initial Training for New Criminal Investigators': {
            'NA': 'Criminal_New_Employee_Training',
            'Training on New Laws Related to the Worst Forms of Child Labor': 'Criminal_New_Law_Training',
            'Refresher Courses Provided': 'Criminal_Refresher_Courses'
        },
        'Reciprocal Referral Mechanism Exists Between Criminal Authorities and Social Services':
            'Criminal_Referral_Mechanism',
        'Imposed Penalties for Violations Related to The Worst Forms of Child Labor': 'Criminal_Penalties_for_WFCL'
    }

    if overview and (overview in tags and tags[overview]):
        if isinstance(tags[overview], dict):
            if sub_overview is None:
                sub_overview = 'NA'
            tag = enforcements.find(tags[overview][sub_overview])
            if tag is None:
                tag = elementTree.SubElement(
                    enforcements, tags[overview][sub_overview])
        else:
            tag = enforcements.find(tags[overview])
            if tag is None:
                tag = elementTree.SubElement(
                    enforcements, tags[overview])
        if multiple_territories:
            tag = elementTree.SubElement(tag, 'Territory')
            display_name = 'All Territories'
            if related_entity:
                display_name = country_display_names[related_entity] \
                    if related_entity in country_display_names else related_entity
            elementTree.SubElement(tag, 'Territory_Name').text = display_name
            elementTree.SubElement(tag, 'Territory_Display_Name').text = display_name
            elementTree.SubElement(tag, 'Enforcement').text = current_year_data
        else:
            tag.text = current_year_data


def government_actions(country, row):
    actions = country.find('Suggested_Actions')
    if actions is None:
        actions = elementTree.SubElement(country, 'Suggested_Actions')

    data = transform_data_to_list(row)

    area = data[3]
    action = data[4].replace('\n', ' ')

    tags = {
        'Legal Framework': 'Legal_Framework',
        'Enforcement': 'Enforcement',
        'Coordination': 'Coordination',
        'Government Policies': 'Government_Policies',
        'Social Programs': 'Social_Programs'
    }

    if area and action:
        if area in tags:
            tag = actions.find(tags[area])
            if tag is None:
                tag = elementTree.SubElement(actions, tags[area])

            action_tag = elementTree.SubElement(tag, 'Action')
            elementTree.SubElement(action_tag, 'Name').text = action


def deliberative_data(country, row):
    mechanisms = country.find('Mechanisms')
    if mechanisms is None:
        mechanisms = elementTree.SubElement(country, 'Mechanisms')

    data = transform_data_to_list(row)

    yes_no_na = data[2]
    program = data[3]

    tags = {
        'Does the government have a program to combat child labor?': 'Program',
        'Does the government have a policy to combat child labor?': 'Policy',
        'Does the government have a mechanism to coordinate efforts against CL?': 'Coordination'
    }

    if program and program in tags:
        elementTree.SubElement(mechanisms, tags[program]).text = yes_no_na


def read_row(country, row, ws_idx):
    options = {
        1: country_profiles,
        2: statistics_on_children,
        3: ratification_of_international,
        4: laws_and_regulations,
        5: labor_law_enforcement,
        6: criminal_law_enforcement,
        7: government_actions,
        8: deliberative_data,
        9: goods_list
        # 11: labor_inspector_info
    }

    data = transform_data_to_list(row)

    if ws_idx in options:
        if ws_idx == 9 or str(data[0]) == year:
            options[ws_idx](country, row)


def get_countries_key(elem):
    return elem.findtext('Name')


def get_goods_key(elem):
    return elem.findtext('Good_Name')


def indent(elem, level=0):
    i = '\n' + level*'  '
    if len(elem):
        if elem.text is None or not elem.text.strip():
            elem.text = i + '  '
        if elem.tail is None or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level+1)
        if elem.tail is None or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
# sanitize_re = re.compile(r"\(|\)|,|the")


def sanitize(text):
    # text = re.sub(sanitize_re, '', text)
    if text is None:
        return
    chars_to_replace = ['(', ')', ',', 'the']
    for char in chars_to_replace:
        text.replace(char, '')
    text = text.replace('  ', ' ')
    text = text.lower().replace(' ', '-')
    return text


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


skip = ['Instructions']

# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *
do_not_read = {
    'Instructions': 1,
    'Country Profiles': 0,
    '1.Statistics on Children': 0,
    '3.Ratification of International': 0,
    '4.Laws and Regulations': 0,
    '6.Labor Law Enforcement Efforts': 0,
    '7.Criminal Law Enforcement Effo': 0,
    '11.Suggested Government Actions': 0,
    'Table H App Deliberative Data': 0,
    'goods_list': 0
}

for index, sheet in enumerate(wb.sheetnames):
    if do_not_read[sheet]:
        continue
    ws = wb[sheet]

    n = 0
    print(sheet)

    for it_row in ws.iter_rows(min_row=2, values_only=True):
        ws_country_name = it_row[1]
        ws_country = country_exists(ws_country_name)

        if type(ws_country) != elementTree.Element and ws_country_name is not None:
            ws_country = elementTree.SubElement(countries, 'Country')
            ws_name = elementTree.SubElement(ws_country, 'Name')
            ws_name.text = ws_country_name
            elementTree.SubElement(ws_country, 'Webpage').text = \
                'https://www.dol.gov/agencies/ilab/resources/reports/child-labor/' + sanitize(ws_country_name)

        read_row(ws_country, it_row, index)

countries[:] = sorted(countries, key=get_countries_key)
indent(countries)
countries_tree.write(COUNTRIES_OUTPUT_FILE, encoding='utf-8')

goods[:] = sorted(goods, key=get_goods_key)
indent(goods)
goods_tree.write(GOODS_OUTPUT_FILE)
